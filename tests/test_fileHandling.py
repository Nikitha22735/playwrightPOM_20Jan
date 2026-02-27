import csv
import json
import os

from dotenv import load_dotenv
from openpyxl import load_workbook
from playwright.sync_api import Page, expect
import pytest
from pages.loginPage import loginClass
from utils.handlingCSV import readCSV
from utils.handlingJsonData import readJsonData



@pytest.mark.login()
def test_example_JsonHandling(page: Page, launchAmazon):
    page.wait_for_timeout(5000)
    login = loginClass(page)
    login.hoverOnAccountAndList()
    # login.clickOnSigInBtn()
    credentials = readJsonData("testData\\credentials.json")
    print(credentials["positiveCredentials"]["username"])
    login.enteruserName(credentials["positiveCredentials"]["username"])
    login.clickOnContinueBtn()
    login.enterPassword(credentials["positiveCredentials"]["password"])
    login.clickOnLogInBtn()



def test_example_readingCSVFile(page: Page, launchAmazon):
    page.wait_for_timeout(5000)
    login = loginClass(page)
    login.hoverOnAccountAndList()
    # login.clickOnSigInBtn()
    credentials = []
    credentials = readCSV("testData\\credentails.csv")
    print(credentials)
    print(credentials[0]["username"])


def test_example_cml(page: Page, launchAmazon):
    page.wait_for_timeout(5000)
    login = loginClass(page)
    login.hoverOnAccountAndList()
    # login.clickOnSigInBtn()
    login.enteruserName(os.getenv("usname"))
    login.clickOnContinueBtn()
    login.enterPassword(os.getenv("pw"))
    login.clickOnLogInBtn()


def test_example_envFile(page: Page, launchAmazon):
    page.wait_for_timeout(5000)
    login = loginClass(page)
    login.hoverOnAccountAndList()
    # login.clickOnSigInBtn()
    load_dotenv(".env")
    login.enteruserName(os.getenv("usname"))
    login.clickOnContinueBtn()
    login.enterPassword(os.getenv("pw"))
    login.clickOnLogInBtn()




# @pytest.mark.parametrize("env",[".env.dev",".env.test",".env.prod"])
def test_example_envFile(page: Page, launchAmazon):
    page.wait_for_timeout(5000)
    login = loginClass(page)
    login.hoverOnAccountAndList()
    # login.clickOnSigInBtn()
    env = os.getenv("env")
    # load_dotenv(f".env.{env}")
    load_dotenv(os.getenv("env", ".env.dev"))
    # print(env)
    # load_dotenv(env)
    print("=============================")
    print(os.getenv("url", "https://www.amazon.in/?"))
    print(os.getenv("usname","trainingplaywright@gmail.com"))
    login.enterPassword("Welcome@04")

@pytest.mark.env()
 # pip install openpyxl
def test_example_excelFile():
    workbook = load_workbook("testData\\searchProduct.xlsx")
    sheetValues = workbook["phoneConfigs"]
    # data = []
    # for row in sheetValues.iter_cols(min_col=1,max_col=2,values_only=True):
    #     data.append(row)


    print(sheetValues["A5"].value)
    

   






